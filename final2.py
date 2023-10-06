import random
import string
import pyperclip
import time
import openpyxl
from openpyxl import Workbook
import os

# List of Hindi words transliterated to English (you can extend this list)
hindi_words_transliterated = [
    "Sundar", "Sukha", "Pyaara", "Chhota", "Mota", "Bura", "Khoobsurat",
    "Swasth", "Garib", "Ameer", "Shant", "Jhulta", "Majedar", "Garam",
    "Thanda", "Naya", "Purana", "Lal", "Hara", "Neela", "Safed", "Kala",
    "Sukhi", "Dhaniya", "Lamba", "Samridh", "Sachcha", "Jhoota", "Sharif",
    "Tez", "Teekha", "Udta", "Langda", "Jhagdalu", "Budbak", "Gaddar", "Bhutiya"
]

hindi_words_transliterated2 = [
    "Kanya",  # Love (Noun)
    "Talwar",  # Happiness (Noun)
    "Teer",  # Journey (Noun)
    "Tyagi",  # Heart (Noun)
    "Chhatri",  # Weather (Noun)
    "Maal",  # Life (Noun)
    "Rasta",  # Roads (Noun)
    "Mazdur",  # World (Noun)
    "kitab",  # Book (Noun)
    "kaagaz",  # Paper (Noun)
    "kursi",  # Chair (Noun)
    "botal",  # Bottle (Noun)
    "raja",  # King (Gangster Name)
    "don",  # Don (Gangster Name)
    "bhai",  # Brother (Gangster Name)
    "Shakeel",  # The Biggest (Gangster Name)
    "gunda",  # Goon (Gangster Name)
    "Aurat",  # Lioness (Gangster Name)
    "Panchi",  # God (Gangster Name)
    "Khidki", "Darwaza", "Mahal", "Gulaab"
]


def replace_characters(word):
    # Replace 's' with '$', 'o' with '0'
    word = word.replace('s', '$')
    word = word.replace('o', '0')
    return word


def generate_strong_password(password_length):
    if password_length < 6:
        return "Password length is too short (minimum length is 6 characters)."

    while True:
        word1 = random.choice(hindi_words_transliterated)
        word2 = random.choice(hindi_words_transliterated2)
        if word1 != word2:
            break

    word1 = replace_characters(word1).capitalize()
    word2 = replace_characters(word2).capitalize()

    digit = random.choice(string.digits)
    special_char = random.choice(string.punctuation)

    password = [word1, word2, digit, special_char]

    indicestoshuffle = [0, 2, 3]
    shuffled_elements = random.sample([password[i] for i in indicestoshuffle], len(indicestoshuffle))
    random.shuffle(shuffled_elements)
    for i, index in enumerate(indicestoshuffle):
        password[index] = shuffled_elements[i]

    password = ''.join(map(str, password))
    readable = f"{word1} {word2}"

    return password, readable


def generate_passwords(num_passwords=5):
    password_length = 11
    passwords = []
    readables = []

    for _ in range(num_passwords):
        password, readable = generate_strong_password(password_length)
        passwords.append(password)
        readables.append(readable)

    for i, each in enumerate(passwords):
        print(f"Generated Strong Password: {i + 1}. {readables[i]} - {each}")

    return passwords

def CreateExcel(rowdata):
    # Check if the Excel file exists
    os.chdir(r"G:\PY\passowordgen")
    file_path = 'PasswordRecords.xlsx' # Use the same file_path variable
    if os.path.exists(file_path):
        # File exists, open it
        workbook = openpyxl.load_workbook(file_path)
        # print(f"Opening existing Excel file: {file_path}")
    else:
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        workbook.save(file_path)

    # Create a new worksheet (if needed)
    worksheet = workbook.active

    # Check if the first row is already filled
    first_row_empty = all(cell.value is None for cell in worksheet[1])

    if first_row_empty:
        # Add your header data to the first row
        header = ["Password","Application_Account"]
        for i, value in enumerate(header, start=1):
            worksheet.cell(row=1, column=i, value=value)
        worksheet.append(rowdata)
    else:
        # Add your data to the first row
        worksheet.append(rowdata)

    workbook.save("PasswordRecords.xlsx")
    workbook.close()
    return

def main():
    try:
        while True:
            passwords = generate_passwords()
            print("")
            choice = int(input("Press Number to Copy or 11 to Regenerate: "))
            if choice != 11:
                if 1 <= choice <= len(passwords):
                    pyperclip.copy(passwords[choice - 1])
                    copied_text = pyperclip.paste()
                    if copied_text == passwords[choice - 1]:
                        save_choice = input("What app or account do you want to save this with: ")
                        row = [passwords[choice - 1], save_choice]
                        CreateExcel(row)
                        print(f"{row} - Added to the Passwords Records")
                        print("Password copied to clipboard!")
                        print("Thankyou For Using AnanyPass")
                        time.sleep(3)
                    else:
                        print("Copying to clipboard failed.")
                    break
                else:
                    print("Invalid choice. Please choose a valid option.")
            elif choice == 11:
                pass
    except ValueError:
        print('Exiting')


if __name__ == "__main__":
    main()
