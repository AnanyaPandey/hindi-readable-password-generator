import openpyxl
import os

# Load the Excel workbook
os.chdir(r"G:\py\passowordgen")
file_path = 'PasswordRecords.xlsx'  # Replace with the path to your Excel file

workbook = openpyxl.load_workbook(file_path)

# Select the active worksheet
worksheet = workbook.active

# Check if the first row is empty
first_row_empty = all(cell.value is None for cell in worksheet[1])

if first_row_empty:
    print("The first row is empty.")
    print("Creating Headers")
    # Add your header data to the first row
    header = ["First Name", "Last Name", "Age"]
    for i, value in enumerate(header, start=1):
        worksheet.cell(row=1, column=i, value=value)
else:
    print("The first row is not empty.")
    print("Added the Data")
    # Add your data to the first row
    data = ["Ananya", "Pandey", 34]
    #for i, value in enumerate(data, start=2):
        #worksheet.cell(row=i, column=i, value=value)
    worksheet.append(data)
workbook.save("PasswordRecords.xlsx")
# Close the workbook
workbook.close()
