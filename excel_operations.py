import os
import openpyxl

def CreateExcel(rowdata):
    # ... (same as in your original code)
    # Check if the Excel file exists
    os.chdir(r"G:\PY\passowordgen")
    file_path = 'PasswordRecords.xlsx' # Use the same file_path variable
    if os.path.exists(file_path):
        # File exists, open it
        workbook = openpyxl.load_workbook(file_path)
        # print(f"Opening existing Excel file: {file_path}")
    else:
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        workbook.save(file_path)

    # Create a new worksheet (if needed)
    worksheet = workbook.active

    # Check if the first row is already filled
    first_row_empty = all(cell.value is None for cell in worksheet[1])

    if first_row_empty:
        # Add your header data to the first row
        header = ["Password","Application_Account"]
        for i, value in enumerate(header, start=1):
            worksheet.cell(row=1, column=i, value=value)
        worksheet.append(rowdata)
    else:
        # Add your data to the first row
        worksheet.append(rowdata)

    workbook.save("PasswordRecords.xlsx")
    workbook.close()
    return
